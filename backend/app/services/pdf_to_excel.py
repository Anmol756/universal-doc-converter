"""
PDF → Excel conversion service.
Uses pdfplumber to extract tables and text from PDF,
then openpyxl to write them into an Excel workbook.
"""

import logging
from pathlib import Path

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logger = logging.getLogger(__name__)


def convert_pdf_to_excel(input_path: Path, output_path: Path) -> Path:
    """
    Convert a PDF file to Excel (XLSX) format.

    Extracts tables from each page using pdfplumber.
    If no tables are found, extracts text content line by line.

    Args:
        input_path: Path to the source PDF file
        output_path: Path where the XLSX file will be saved

    Returns:
        Path to the converted XLSX file
    """
    if not input_path.exists():
        raise ValueError(f"Input file not found: {input_path}")

    if input_path.suffix.lower() != ".pdf":
        raise ValueError(f"Expected PDF file, got: {input_path.suffix}")

    output_path = output_path.with_suffix(".xlsx")

    logger.info(f"Converting PDF → Excel: {input_path.name} → {output_path.name}")

    try:
        wb = Workbook()
        # Remove default sheet — we'll create named sheets
        wb.remove(wb.active)

        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_alignment = Alignment(vertical="top", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        with pdfplumber.open(str(input_path)) as pdf:
            total_tables = 0

            for page_num, page in enumerate(pdf.pages):
                tables = page.extract_tables()

                if tables:
                    for table_idx, table in enumerate(tables):
                        total_tables += 1
                        sheet_name = f"Page{page_num + 1}_Table{table_idx + 1}"
                        # Excel sheet names max 31 chars
                        sheet_name = sheet_name[:31]
                        ws = wb.create_sheet(title=sheet_name)

                        for row_idx, row in enumerate(table):
                            for col_idx, cell_value in enumerate(row):
                                cell = ws.cell(
                                    row=row_idx + 1,
                                    column=col_idx + 1,
                                    value=cell_value or "",
                                )
                                cell.alignment = cell_alignment
                                cell.border = thin_border

                                # Style first row as header
                                if row_idx == 0:
                                    cell.font = header_font
                                    cell.fill = header_fill
                                    cell.alignment = header_alignment

                        # Auto-fit column widths (approximate)
                        for col_idx in range(len(table[0]) if table else 0):
                            max_len = max(
                                (len(str(row[col_idx] or "")) for row in table),
                                default=10,
                            )
                            ws.column_dimensions[chr(65 + col_idx)].width = min(max_len + 4, 50)

            # If no tables found, extract text content
            if total_tables == 0:
                ws = wb.create_sheet(title="Text Content")
                ws.cell(row=1, column=1, value="Page").font = header_font
                ws.cell(row=1, column=1).fill = header_fill
                ws.cell(row=1, column=2, value="Content").font = header_font
                ws.cell(row=1, column=2).fill = header_fill
                ws.column_dimensions["A"].width = 10
                ws.column_dimensions["B"].width = 100

                row_num = 2
                for page_num, page in enumerate(pdf.pages):
                    text = page.extract_text()
                    if text:
                        for line in text.split("\n"):
                            line = line.strip()
                            if line:
                                ws.cell(row=row_num, column=1, value=page_num + 1)
                                ws.cell(row=row_num, column=2, value=line)
                                ws.cell(row=row_num, column=2).alignment = cell_alignment
                                row_num += 1

        # Ensure at least one sheet exists
        if len(wb.sheetnames) == 0:
            ws = wb.create_sheet(title="Empty")
            ws.cell(row=1, column=1, value="No content extracted from PDF")

        wb.save(str(output_path))

    except Exception as e:
        logger.error(f"PDF to Excel conversion failed: {e}")
        raise RuntimeError(f"Conversion failed: {str(e)}") from e

    if not output_path.exists():
        raise RuntimeError("Conversion produced no output file")

    logger.info(f"Conversion complete: {output_path.name} ({output_path.stat().st_size} bytes)")
    return output_path
