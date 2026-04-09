"""
Excel → PDF conversion service.
Uses openpyxl to read the spreadsheet and reportlab to generate a PDF.
Renders worksheet data as formatted tables in the PDF.
"""

import logging
from pathlib import Path

from openpyxl import load_workbook
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak,
)

logger = logging.getLogger(__name__)


def convert_excel_to_pdf(input_path: Path, output_path: Path) -> Path:
    """
    Convert an Excel (XLSX) file to PDF format.

    Reads each worksheet and renders the data as styled tables
    in the resulting PDF document.

    Args:
        input_path: Path to the source XLSX file
        output_path: Path where the PDF file will be saved

    Returns:
        Path to the converted PDF file
    """
    if not input_path.exists():
        raise ValueError(f"Input file not found: {input_path}")

    if input_path.suffix.lower() not in (".xlsx", ".xls"):
        raise ValueError(f"Expected Excel file, got: {input_path.suffix}")

    output_path = output_path.with_suffix(".pdf")

    logger.info(f"Converting Excel → PDF: {input_path.name} → {output_path.name}")

    try:
        wb = load_workbook(str(input_path), data_only=True)

        pdf = SimpleDocTemplate(
            str(output_path),
            pagesize=landscape(A4),
            rightMargin=36,
            leftMargin=36,
            topMargin=36,
            bottomMargin=36,
        )

        styles = getSampleStyleSheet()
        sheet_title_style = ParagraphStyle(
            "SheetTitle",
            parent=styles["Heading2"],
            fontSize=14,
            spaceAfter=12,
            fontName="Helvetica-Bold",
        )

        story = []

        for sheet_idx, sheet_name in enumerate(wb.sheetnames):
            ws = wb[sheet_name]

            # Sheet title
            story.append(Paragraph(f"Sheet: {sheet_name}", sheet_title_style))

            # Extract data
            data = []
            for row in ws.iter_rows(values_only=True):
                cleaned_row = [str(cell) if cell is not None else "" for cell in row]
                data.append(cleaned_row)

            if not data:
                story.append(Paragraph("<i>Empty sheet</i>", styles["Normal"]))
            else:
                # Calculate column widths (distribute evenly across page width)
                num_cols = max(len(row) for row in data) if data else 1
                available_width = landscape(A4)[0] - 72  # minus margins
                col_width = min(available_width / num_cols, 200)
                col_widths = [col_width] * num_cols

                table = Table(data, colWidths=col_widths, repeatRows=1)

                table.setStyle(TableStyle([
                    # Header row
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2B579A")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 9),
                    ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
                    ("TOPPADDING", (0, 0), (-1, 0), 8),

                    # Data rows
                    ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                    ("FONTSIZE", (0, 1), (-1, -1), 8),
                    ("ALIGN", (0, 1), (-1, -1), "LEFT"),
                    ("TOPPADDING", (0, 1), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 1), (-1, -1), 4),

                    # Grid
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),

                    # Alternating row colors
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1),
                     [colors.white, colors.HexColor("#F5F5F5")]),
                ]))

                story.append(table)

            # Page break between sheets
            if sheet_idx < len(wb.sheetnames) - 1:
                story.append(PageBreak())

        if not story:
            story.append(Paragraph("(Empty workbook)", styles["Normal"]))

        pdf.build(story)

    except Exception as e:
        logger.error(f"Excel to PDF conversion failed: {e}")
        raise RuntimeError(f"Conversion failed: {str(e)}") from e

    if not output_path.exists():
        raise RuntimeError("Conversion produced no output file")

    logger.info(f"Conversion complete: {output_path.name} ({output_path.stat().st_size} bytes)")
    return output_path
